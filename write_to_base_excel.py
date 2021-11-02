import openpyxl
import shutil

# 엑셀 템플릿 파일 경로
base_dir_path = '/home/hong/Documents/pythonProject/write_to_base_excel/'
base_year_path = 'base_year.xlsx'
base_month_path = 'base_month.xlsx'
base_day_path = 'base_day.xlsx'
# load_data_path = 'tmy_era_47.1_15.197_2005_2014.csv'

# 다른이름으로 저장(ex: test_year.xlsx)
save_file_name = 'test'
write_file_year = base_dir_path + save_file_name + '_year.xlsx'
write_file_month = base_dir_path + save_file_name + '_month.xlsx'
write_file_day = base_dir_path + save_file_name + '_day.xlsx'

# 템플릿 엑셀을 다른이름 엑셀파일로 복사
shutil.copy(base_dir_path + base_year_path, write_file_year)
shutil.copy(base_dir_path + base_month_path, write_file_month)
shutil.copy(base_dir_path + base_day_path, write_file_day)

# 복사한 파일 로드(쓰기 위한 준비)
year_wb = openpyxl.load_workbook(write_file_year)
month_wb = openpyxl.load_workbook(write_file_month)
day_wb = openpyxl.load_workbook(write_file_day)

# 매개변수
year = 2021
month = 11
month_days_count = 30
day = 2
location = '판교'

# 더미 데이터
temperature, humidity, wind_direction, wind_speed, rain, global_horizontal_irr, diffuse_horizontal_irr, direct_normal_irr\
    = 1, 1, 1, 1, 1, 1, 1, 1

########################################################################################################################
# 연보양식
temp_sum, hum_sum, wd_sum, ws_sum, rain_sum, ghi_sum, dhi_sum, dni_sum = 0,0,0,0,0,0,0,0
ws = year_wb.active
ws['A2'] = '{}년'.format(year)
ws['J2'] = location
for i in range(1,13):
    ws.cell(7+i, 1, i)
    ws.cell(7+i, 2, temperature)
    ws.cell(7 + i, 3, humidity)
    ws.cell(7 + i, 4, wind_direction)
    ws.cell(7 + i, 5, wind_speed)
    ws.cell(7 + i, 6, rain)
    ws.cell(7 + i, 7, global_horizontal_irr)
    ws.cell(7 + i, 8, diffuse_horizontal_irr)
    ws.cell(7 + i, 9, direct_normal_irr)
    temp_sum += temperature
    hum_sum += humidity
    wd_sum += wind_direction
    ws_sum += wind_speed
    rain_sum += rain
    ghi_sum += global_horizontal_irr
    dhi_sum += diffuse_horizontal_irr
    dni_sum += direct_normal_irr
    
# 평균
ws['B39'], ws['C39'], ws['D39'], ws['E39'], ws['F39'], ws['G39'], ws['H39'], ws['I39'] = \
    temp_sum/12, hum_sum/12, wd_sum/12, ws_sum/12, rain_sum/12, ghi_sum/12, dhi_sum/12, dni_sum/12
# 합계
ws['B40'], ws['C40'], ws['D40'], ws['E40'], ws['F40'], ws['G40'], ws['H40'], ws['I40'] = \
    temp_sum, hum_sum, wd_sum, ws_sum, rain_sum, ghi_sum, dhi_sum, dni_sum
year_wb.save(write_file_year)

########################################################################################################################
# 월보양식
temp_sum, hum_sum, wd_sum, ws_sum, rain_sum, ghi_sum, dhi_sum, dni_sum = 0,0,0,0,0,0,0,0
ws = month_wb.active
ws['A2'] = '{}년-{}월'.format(year, month)
ws['J2'] = location

for i in range(1, month_days_count+1):
    ws.cell(7 + i, 1, '{}/{}'.format(month, i))
    ws.cell(7 + i, 2, temperature)
    ws.cell(7 + i, 3, humidity)
    ws.cell(7 + i, 4, wind_direction)
    ws.cell(7 + i, 5, wind_speed)
    ws.cell(7 + i, 6, rain)
    ws.cell(7 + i, 7, global_horizontal_irr)
    ws.cell(7 + i, 8, diffuse_horizontal_irr)
    ws.cell(7 + i, 9, direct_normal_irr)
    temp_sum += temperature
    hum_sum += humidity
    wd_sum += wind_direction
    ws_sum += wind_speed
    rain_sum += rain
    ghi_sum += global_horizontal_irr
    dhi_sum += diffuse_horizontal_irr
    dni_sum += direct_normal_irr
    
# 평균
ws['B39'], ws['C39'], ws['D39'], ws['E39'], ws['F39'], ws['G39'], ws['H39'], ws['I39'] = \
    temp_sum/month_days_count, hum_sum/month_days_count, wd_sum/month_days_count, ws_sum/month_days_count, \
    rain_sum/month_days_count, ghi_sum/month_days_count, dhi_sum/month_days_count, dni_sum/month_days_count
# 합계
ws['B40'], ws['C40'], ws['D40'], ws['E40'], ws['F40'], ws['G40'], ws['H40'], ws['I40'] = \
    temp_sum, hum_sum, wd_sum, ws_sum, rain_sum, ghi_sum, dhi_sum, dni_sum
month_wb.save(write_file_month)

########################################################################################################################
# 일보양식
temp_sum, hum_sum, wd_sum, ws_sum, rain_sum, ghi_sum, dhi_sum, dni_sum = 0,0,0,0,0,0,0,0
ws = day_wb.active
ws['A2'] = '{}년-{}월-{}일'.format(year, month, day)
ws['J2'] = location
ws['A58'] = '{}년-{}월-{}일'.format(year, month, day)
ws['J58'] = location
ws['A113'] = '{}년-{}월-{}일'.format(year, month, day)
ws['J113'] = location
time = 0

for i in range(8, 57):
    ws.cell(i, 1, ('%d:%02d' % (int(time / 60), int(time % 60))))
    ws.cell(i, 2, temperature)
    ws.cell(i, 3, humidity)
    ws.cell(i, 4, wind_direction)
    ws.cell(i, 5, wind_speed)
    ws.cell(i, 6, rain)
    ws.cell(i, 7, global_horizontal_irr)
    ws.cell(i, 8, diffuse_horizontal_irr)
    ws.cell(i, 9, direct_normal_irr)
    time += 10
    temp_sum += temperature
    hum_sum += humidity
    wd_sum += wind_direction
    ws_sum += wind_speed
    rain_sum += rain
    ghi_sum += global_horizontal_irr
    dhi_sum += diffuse_horizontal_irr
    dni_sum += direct_normal_irr

for i in range(64, 112):
    ws.cell(i, 1, ('%d:%02d' % (int(time / 60), int(time % 60))))
    ws.cell(i, 2, temperature)
    ws.cell(i, 3, humidity)
    ws.cell(i, 4, wind_direction)
    ws.cell(i, 5, wind_speed)
    ws.cell(i, 6, rain)
    ws.cell(i, 7, global_horizontal_irr)
    ws.cell(i, 8, diffuse_horizontal_irr)
    ws.cell(i, 9, direct_normal_irr)
    time += 10
    temp_sum += temperature
    hum_sum += humidity
    wd_sum += wind_direction
    ws_sum += wind_speed
    rain_sum += rain
    ghi_sum += global_horizontal_irr
    dhi_sum += diffuse_horizontal_irr
    dni_sum += direct_normal_irr

for i in range(119, 166):
    ws.cell(i, 1, ('%d:%02d' % (int(time / 60), int(time % 60))))
    ws.cell(i, 2, temperature)
    ws.cell(i, 3, humidity)
    ws.cell(i, 4, wind_direction)
    ws.cell(i, 5, wind_speed)
    ws.cell(i, 6, rain)
    ws.cell(i, 7, global_horizontal_irr)
    ws.cell(i, 8, diffuse_horizontal_irr)
    ws.cell(i, 9, direct_normal_irr)
    time += 10
    temp_sum += temperature
    hum_sum += humidity
    wd_sum += wind_direction
    ws_sum += wind_speed
    rain_sum += rain
    ghi_sum += global_horizontal_irr
    dhi_sum += diffuse_horizontal_irr
    dni_sum += direct_normal_irr

# 평균
ws['B166'], ws['C166'], ws['D166'], ws['E166'], ws['F166'], ws['G166'], ws['H166'], ws['I166'] = \
    temp_sum/144, hum_sum/144, wd_sum/144, ws_sum/144, rain_sum/144, ghi_sum/144, dhi_sum/144, dni_sum/144
# 합계
ws['B167'], ws['C167'], ws['D167'], ws['E167'], ws['F167'], ws['G167'], ws['H167'], ws['I167'] = \
    temp_sum, hum_sum, wd_sum, ws_sum, rain_sum, ghi_sum, dhi_sum, dni_sum
day_wb.save(write_file_day)

